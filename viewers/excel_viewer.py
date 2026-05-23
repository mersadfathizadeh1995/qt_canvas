"""Excel viewer — openpyxl read-only with sheet picker.

Skips registration when openpyxl is not installed; files of that
type then fall through to the binary viewer.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, List, Optional, Tuple

from PySide6.QtCore import QAbstractTableModel, QModelIndex, Qt
from PySide6.QtWidgets import (
    QComboBox,
    QHBoxLayout,
    QLabel,
    QTableView,
    QVBoxLayout,
    QWidget,
)

from ..registry import ViewerSpec, register_viewer
from .base import FileViewer

log = logging.getLogger(__name__)

#: Row cap per sheet.  openpyxl read-only can stream more, but
#: rendering more than this in a QTableView gets unresponsive.
ROW_CAP: int = 5000

#: Column cap per sheet — same rationale.
COLUMN_CAP: int = 200

#: Hard file-size cap; above this :class:`ViewerStack` short-circuits
#: to the binary viewer.
MAX_EXCEL_BYTES: int = 100 * 1024 * 1024  # 100 MB

try:
    import openpyxl as _openpyxl  # type: ignore
    _OPENPYXL_OK = True
except Exception:  # noqa: BLE001
    _OPENPYXL_OK = False


class _ExcelTableModel(QAbstractTableModel):
    def __init__(
        self,
        headers: List[str],
        rows: List[List[Any]],
        parent: Optional[QWidget] = None,
    ) -> None:
        super().__init__(parent)
        self._headers = list(headers)
        self._rows = list(rows)

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: N802
        return 0 if parent.isValid() else len(self._rows)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: N802
        return 0 if parent.isValid() else len(self._headers)

    def data(self, index: QModelIndex, role: int = Qt.DisplayRole) -> Any:
        if not index.isValid() or role != Qt.DisplayRole:
            return None
        row = self._rows[index.row()]
        col = index.column()
        if col >= len(row):
            return ""
        value = row[col]
        return "" if value is None else str(value)

    def headerData(  # noqa: N802
        self,
        section: int,
        orientation: Qt.Orientation,
        role: int = Qt.DisplayRole,
    ) -> Any:
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return self._headers[section] if 0 <= section < len(self._headers) else ""
        return section + 1


class ExcelViewer(FileViewer):
    """Excel preview backed by ``openpyxl`` (read-only)."""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        controls = QHBoxLayout()
        controls.setContentsMargins(0, 0, 0, 0)
        self._sheet_label = QLabel("Sheet:", self)
        self._sheet_picker = QComboBox(self)
        self._sheet_picker.currentTextChanged.connect(self._on_sheet_changed)
        controls.addWidget(self._sheet_label)
        controls.addWidget(self._sheet_picker, 1)
        layout.addLayout(controls)

        self._view = QTableView(self)
        self._view.setAlternatingRowColors(True)
        layout.addWidget(self._view, 1)

        self._workbook: Optional[Any] = None
        self._path: Optional[Path] = None
        self._model: Optional[_ExcelTableModel] = None

    # ------------------------------------------------------------------
    def load_path(self, path: Path) -> None:
        try:
            wb = _openpyxl.load_workbook(
                filename=str(path), read_only=True, data_only=True
            )
        except Exception as exc:  # noqa: BLE001
            self.load_failed.emit(path, f"openpyxl failed: {exc}")
            return
        self._path = path
        self._workbook = wb
        self._sheet_picker.blockSignals(True)
        self._sheet_picker.clear()
        self._sheet_picker.addItems(list(wb.sheetnames))
        self._sheet_picker.blockSignals(False)
        if wb.sheetnames:
            self._load_sheet(wb.sheetnames[0])
        self.file_loaded.emit(path)

    def clear(self) -> None:
        self._sheet_picker.clear()
        self._view.setModel(None)
        if self._workbook is not None:
            try:
                self._workbook.close()
            except Exception:  # noqa: BLE001
                pass
        self._workbook = None
        self._path = None
        self._model = None

    def max_preview_bytes(self) -> Optional[int]:
        return MAX_EXCEL_BYTES

    # ------------------------------------------------------------------
    def _on_sheet_changed(self, name: str) -> None:
        if not name or self._workbook is None:
            return
        self._load_sheet(name)

    def _load_sheet(self, name: str) -> None:
        ws = self._workbook[name]
        headers, rows = self._read_sheet(ws)
        self._model = _ExcelTableModel(headers, rows, self)
        self._view.setModel(self._model)
        for col in range(self._model.columnCount()):
            self._view.resizeColumnToContents(col)

    @staticmethod
    def _read_sheet(ws) -> Tuple[List[str], List[List[Any]]]:
        rows: List[List[Any]] = []
        headers: List[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [
                    (str(c) if c is not None else f"col_{j + 1}")
                    for j, c in enumerate(row[:COLUMN_CAP])
                ]
                continue
            rows.append(list(row[:COLUMN_CAP]))
            if len(rows) >= ROW_CAP:
                break
        return headers, rows


if _OPENPYXL_OK:
    register_viewer(ViewerSpec(
        key="excel",
        label="Excel",
        extensions=(".xlsx", ".xlsm"),
        factory=ExcelViewer,
        priority=10,
        description="Excel preview (read-only) via openpyxl.",
    ))
else:
    log.debug("openpyxl missing — excel viewer not registered (falls back to binary)")


__all__ = ["ExcelViewer", "MAX_EXCEL_BYTES", "ROW_CAP", "COLUMN_CAP"]
